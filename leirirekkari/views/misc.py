# -*- coding: utf-8 -*-
from pyramid.response import Response
from pyramid.view import view_config
import pyramid.security as security

from sqlalchemy.exc import DBAPIError
from sqlalchemy import or_, and_
from pyramid.httpexceptions import HTTPFound

from leirirekkari.models.dbsession import (
    DBSession
    )

from leirirekkari.models.user import (
    User, Group, Privilege, UserAudit, UserLogin
    )

from leirirekkari.models.setting import Setting
from leirirekkari.models.organization import (
    Club,
    SubUnit,
    Village,
    Subcamp,
    )

from leirirekkari.models.participant import (
    Participant,
    ParticipantPhone,
    ParticipantNextOfKin,
    ParticipantLanguage,
    ParticipantPayment,
    ParticipantWishes,
    ParticipantWishesOption,
    ParticipantSignupOption,
    ParticipantMedical,
    ParticipantMedicalDiet,
    ParticipantMedicalFoodAllergy,
    ParticipantMedicalAllergy,
    ParticipantAddress,
    ParticipantMeta,
    ParticipantPresence,
    ParticipantStatus,
    ParticipantEnlistment,
    ParticipantEnlistmentOption,
    ParticipantPolkuBookings,
    ParticipantPolkuAnswers,
    ParticipantPolkuContactInfo,
    )
    
from leirirekkari.models.medical import (
    MedicalCard,
    MedicalCardEvent,
    MedicalParticipantStatus,
    MedicalParticipantAdditional,
    MedicalReason,
    MedicalTreatmentType,
    MedicalMethodOfArrival,
    )

from leirirekkari.models.setting import (
    Setting,
    Feedback,
    )
import leirirekkari.helpers.helpers as helpers
import leirirekkari.helpers.organization as organizationhelpers
import leirirekkari.helpers.user as userhelpers

from pyramid.i18n import get_localizer, negotiate_locale_name
from pyramid.i18n import TranslationString
from pyramid.i18n import get_locale_name

from leirirekkari import checkBrowser, checkDevice

class MiscViews(object):
    def __init__(self, request):
        if checkBrowser(request) or checkDevice(request):
            request.redirect_forbidden = True
        else:
            request.redirect_forbidden = False
        self.request = request
    
    @view_config(route_name='json_search_participant', renderer='json', permission='view')
    def json_search_participant(self):
        _ = self.request.translate
        if self.request.method == 'POST':
            search_string = self.request.POST.get('searchstr').strip()
            if search_string == '':
                self.request.session.flash(_(u"Empty search, please provide search string."), 'error')
            else:
                participants = DBSession.query(Participant).filter(
                    or_(
                    Participant.firstname.like('%'+search_string+'%'),
                    Participant.lastname.like('%'+search_string+'%'),
                    Participant.nickname.like('%'+search_string+'%'),
                    Participant.member_no.like('%'+search_string+'%')
                    )).all()
                participants_tmp = []
                for participant in participants:
                    participant_tmp = {
                        'id':participant.id,
                        'firstname':helpers.decodeString(participant.firstname),
                        'lastname':helpers.decodeString(participant.lastname),
                    }
                    participants_tmp.append(participant_tmp)
                    
                return json.dumps(participants_tmp)
        return {}
    
    
    @view_config(route_name='feedback_add_box', renderer='feedback/add_box.mak', permission='view')
    def feedback_add_box(self):
        _ = self.request.translate
        
        return {}
        
    @view_config(route_name='feedback_add_box_submit', renderer='feedback/add_box_submit.mak', permission='view')
    def feedback_add_box_submit(self):
        _ = self.request.translate

        if self.request.method == 'POST':
            feedback = Feedback()
            feedback.title = self.request.POST.get('feedback_title')
            feedback.description = self.request.POST.get('feedback_description')
            feedback.type = int(self.request.POST.get('feedback_type'))
            DBSession.add(feedback)
            DBSession.flush()

        return {'success':True}
    
    @view_config(route_name='test_excel_output', renderer='test/excel.mak', permission='view')
    def test_excel_output(self):
        _ = self.request.translate

        from openpyxl import Workbook
        from openpyxl.cell import get_column_letter
        from pyramid.response import FileResponse
        wb = Workbook()

        dest_filename = r'empty_book.xlsx'

        ws = wb.worksheets[0]

        ws.title = "range names"

        for col_idx in xrange(1, 40):
            col = get_column_letter(col_idx)
            for row in xrange(1, 600):
                ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)

        ws = wb.create_sheet()

        ws.title = 'Pi'

        ws.cell('F5').value = 3.14

#        self.request.response.content_type = "application/ms-excel"
#        self.request.response.content_disposition = 'attachment; filename="leirirekkari_test.xls"'
        wb.save(filename = dest_filename)
        return FileResponse(dest_filename, request=self.request)
        #return {'wb':wb}

    @view_config(route_name='robots', renderer='robots.mak', permission=security.NO_PERMISSION_REQUIRED)
    def robots(self):
        return {}
    
def includeme(config):
    config.add_route('json_search_participant', '/json_search/participant/')
    config.add_route('feedback_add_box', '/feedback/add/box/')
    config.add_route('feedback_add_box_submit', '/feedback/add/box/submit/')
    config.add_route('test_excel_output', '/test/excel_output/')
    config.add_route('robots', '/robots.txt/')
