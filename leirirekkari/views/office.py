# -*- coding: utf-8 -*-
from pyramid.response import Response
from pyramid.view import view_config
import pyramid.security as security

from sqlalchemy.exc import DBAPIError
from sqlalchemy import or_, and_
from sqlalchemy import desc
from sqlalchemy import func as safunc
import sqlalchemy.sql as sql
from pyramid.httpexceptions import HTTPFound

from leirirekkari.models.dbsession import (
    DBSession
    )

from leirirekkari.models.user import (
    User, Group, Privilege, UserAudit, UserLogin
    )

from leirirekkari.models.setting import Setting
from leirirekkari.models.organization import (
    Club,
    SubUnit,
    Village,
    VillageKitchen,
    Subcamp,
    )

from leirirekkari.models.participant import (
    Participant,
    ParticipantPhone,
    ParticipantNextOfKin,
    ParticipantLanguage,
    ParticipantPayment,
    ParticipantWishes,
    ParticipantWishesOption,
    ParticipantSignupOption,
    ParticipantMedical,
    ParticipantMedicalDiet,
    ParticipantMedicalFoodAllergy,
    ParticipantMedicalAllergy,
    ParticipantAddress,
    ParticipantMeta,
    ParticipantPresence,
    ParticipantStatus,
    ParticipantEnlistment,
    ParticipantEnlistmentOption,
    ParticipantPolkuBookings,
    ParticipantPolkuAnswers,
    ParticipantPolkuContactInfo,
    )
    
import leirirekkari.helpers.helpers as helpers
import leirirekkari.helpers.organization as organizationhelpers
import leirirekkari.helpers.participant as participanthelpers
import leirirekkari.helpers.user as userhelpers

from leirirekkari.helpers.status import status_key_list


from pyramid.i18n import get_localizer, negotiate_locale_name
from pyramid.i18n import TranslationString
from pyramid.i18n import get_locale_name

from leirirekkari import checkBrowser, checkDevice

import datetime

from openpyxl import Workbook
from openpyxl import style as openpyxl_style
from openpyxl.cell import get_column_letter
from pyramid.response import FileResponse

class OfficeViews(object):
    def __init__(self, request):
        if checkBrowser(request) or checkDevice(request):
            request.redirect_forbidden = True
        else:
            request.redirect_forbidden = False
        self.request = request
    
    @view_config(route_name='office_frontpage', renderer='office/index.mak', permission='office_view')
    def office_frontpage(self):
        if self.request.redirect_forbidden:
            return HTTPFound(location='/forbidden/')
        elif userhelpers.checkUserPasswordChangeNeed(self.request):
            return HTTPFound(location='/settings/me/edit/')
        _ = self.request.translate
        
        to_view = {
            'basic_info': True,
            'booking_no': False,
            'polku_member_id': False,
            'status': True,
            'status_all': False,
            'age_group': True,
            'sex': False,
            'age': False,
            'birthdate': False,
            'club': True,
            'subunit': True,
            'village': True,
            'village_kitchen': False,
            'subcamp': True,
            'presence': True,
            'next_of_kin': False,
            'wishes': False,
            'address': False,
            'phone_and_email': False,
            'languages': False,
            'designation_title': False,
            'designation_all': False,
            'spiritual': False,
            'payments': False,
        }
        
        if security.has_permission('office_participant_view_medical', self.request.context, self.request):
            medical_to_view = {
                'medical_diets': False,
                'medical_diets_boolean': False,
                'medical_food_allergies': False,
                'medical_food_allergies_boolean': False,
                'medical_allergies': False,
                'medical_allergies_boolean': False,
                'medical_other': False,
                'medical_need_help_boolean': False,
                'kitchen_table': False,
            }
            
            to_view = dict(to_view.items() + medical_to_view.items())
        elif security.has_permission('kitchen_view_allergies', self.request.context, self.request):
            medical_to_view = {
                'medical_diets': False,
                'medical_diets_boolean': False,
                'medical_food_allergies': False,
                'medical_food_allergies_boolean': False,
                'medical_allergies': False,
                'medical_allergies_boolean': False,
                'kitchen_table': False,
            }
            to_view = dict(to_view.items() + medical_to_view.items())
        
        self.request.bread.append({'url':'/office/', 'text':_('Office')})
        return {'to_view':to_view}
    
    @view_config(route_name='office_participant_new', renderer='office/participant/new.mak', permission='office_participant_new')
    def office_participant_new(self):
        if self.request.redirect_forbidden:
            return HTTPFound(location='/forbidden/')
        elif userhelpers.checkUserPasswordChangeNeed(self.request):
            return HTTPFound(location='/settings/me/edit/')
        _ = self.request.translate
        
        participant = Participant()
        participant.address_data.append({})
        participant.address_data[0] = ParticipantAddress()
        participant.medical_data = ParticipantMedical()
        participant.next_of_kin_data.append({})
        participant.next_of_kin_data[0] = ParticipantNextOfKin()
        
        participant_id = 0
        
        if self.request.method == 'POST':

            participant.firstname = self.request.POST.get('firstname').strip()
            participant.lastname = self.request.POST.get('lastname').strip()
            participant.nickname = self.request.POST.get('nickname').strip()
            participant.birthdate = helpers.parseFinnishDateFromString(self.request.POST.get('birthdate').strip(), False)
            participant.title = self.request.POST.get('title').strip()
            participant.age_group = self.request.POST.get('age_group').strip()
            member_no = self.request.POST.get('member_no').strip()
            if member_no == '':
                member_no = 0
            elif not member_no.isdigit():
                member_no = 0
            participant.member_no = member_no
            participant.sex = self.request.POST.get('sex').strip()
            participant.spiritual = self.request.POST.get('spiritual').strip()
            participant.club_id = self.request.POST.get('club_id').strip()
            participant.email = self.request.POST.get('email').strip()
            participant.active = 1

            subunit_id = self.request.POST.get('subunit_id').strip()
            if subunit_id > 0:
                participant.subunit_id = subunit_id
                subunit = organizationhelpers.getSubUnit(subunit_id)
                if subunit != None:
                    participant.village_id = subunit.village_id
                    village = organizationhelpers.getVillage(subunit.village_id)
                    if village != None:
                        participant.subcamp_id = village.subcamp_id
            
            if self.request.POST.get('firstname').strip() != '' and self.request.POST.get('lastname').strip() != '':
                DBSession.add(participant)
                DBSession.flush()
                userAudit = UserAudit(self.request.user.id)
                userAudit.model = 'Participant'
                userAudit.model_id = participant.id
                userAudit.action = 'Create'
                userAudit.revision = participant.metadata_revision
                DBSession.add(userAudit)
                DBSession.flush()
                participant_id = participant.id
            
                participant.address_data[0].street = self.request.POST.get('address_street').strip()
                participant.address_data[0].postalcode = self.request.POST.get('address_postalcode').strip()
                participant.address_data[0].city = self.request.POST.get('address_city').strip()
                participant.address_data[0].country = self.request.POST.get('address_country').strip()
            
                participant.next_of_kin_data[0].primary_name = self.request.POST.get('next_of_kin_0_name').strip()
                participant.next_of_kin_data[0].primary_phone = self.request.POST.get('next_of_kin_0_phone').strip()
                participant.next_of_kin_data[0].primary_email = self.request.POST.get('next_of_kin_0_email').strip()
                participant.next_of_kin_data[0].secondary_name = self.request.POST.get('next_of_kin_1_name').strip()
                participant.next_of_kin_data[0].secondary_phone = self.request.POST.get('next_of_kin_1_phone').strip()
                participant.next_of_kin_data[0].secondary_email = self.request.POST.get('next_of_kin_1_email').strip()
            
                if participant_id > 0:
                    participant.address_data[0].participant_id = participant_id
                    DBSession.add(participant.address_data[0])
                    participant.next_of_kin_data[0].participant_id = participant_id
                    DBSession.add(participant.next_of_kin_data[0])
                    DBSession.flush()
            
                if security.has_permission('office_participant_view_medical', self.request.context, self.request):
                    if len(self.request.POST.getall('medical_diet'))>0:
                        participant.medical_data.diets = DBSession.query(ParticipantMedicalDiet).filter(ParticipantMedicalDiet.id.in_(self.request.POST.getall('medical_diet'))).all()
                                
                    if len(self.request.POST.getall('medical_food_allergy'))>0:
                        participant.medical_data.food_allergies = DBSession.query(ParticipantMedicalFoodAllergy).filter(ParticipantMedicalFoodAllergy.id.in_(self.request.POST.getall('medical_food_allergy'))).all()

                    participant.medical_data.additional_food = self.request.POST.get('medical_additional_food').strip()
                    participant.medical_data.drugs_help = int(self.request.POST.get('medical_drugs_help'))
                    participant.medical_data.illnesses = self.request.POST.get('medical_illnesses').strip()

                    if len(self.request.POST.getall('medical_allergy'))>0:
                        participant.medical_data.allergies = DBSession.query(ParticipantMedicalAllergy).filter(ParticipantMedicalAllergy.id.in_(self.request.POST.getall('medical_allergy'))).all()

                    participant.medical_data.additional_health = self.request.POST.get('medical_additional_health').strip()
                    participant.medical_data.week_of_pregnancy = self.request.POST.get('medical_week_of_pregnancy').strip()

                    if participant_id > 0:
                        participant.medical_data.participant_id = participant_id
                        DBSession.add(participant.medical_data)
                        DBSession.flush()
                else:
                    participant.medical_data.participant_id = participant_id
                    DBSession.add(participant.medical_data)
                    DBSession.flush()

                if participant_id > 0:
                    if len(self.request.POST.getall('language')) > 0:
                        languages = self.request.POST.getall('language')
                        for key, lang in enumerate(languages):
                            if lang.strip() != '':
                                language = ParticipantLanguage(participant_id=participant_id, language=lang)
                                DBSession.add(language)
                                DBSession.flush()

                    if len(self.request.POST.getall('phone_number')) > 0:
                        phones = self.request.POST.getall('phone_number')
                        types = self.request.POST.getall('phone_type')
                        for key, phone_num in enumerate(phones):
                            if phone_num.strip() != '':
                                description = types[key]
                                phone = ParticipantPhone(participant_id=participant_id, phone=phone_num, description=description)
                                DBSession.add(phone)
                                DBSession.flush()

                    if len(self.request.POST.getall('presence_starts')) > 0:
                        presence_starts = self.request.POST.getall('presence_starts')
                        presence_ends = self.request.POST.getall('presence_ends')
                        presence_titles = self.request.POST.getall('presence_title')
                        presence_descriptions = self.request.POST.getall('presence_description')
                
                        for key, presence_start in enumerate(presence_starts):
                            if presence_start != '':
                                presence_starts_dt = helpers.parseFinnishDateFromString(presence_start)
                                presence_end = presence_ends[key]
                                presence_ends_dt = None
                                if presence_end != '':
                                    presence_ends_dt = helpers.parseFinnishDateFromString(presence_end)
                                presence_title = presence_titles[key]
                                presence_description = presence_descriptions[key]
                                presence = ParticipantPresence(participant_id=participant_id, starts = presence_starts_dt, ends = presence_ends_dt, title = presence_title, description = presence_description)
                                DBSession.add(presence)
                                DBSession.flush()

                    if len(self.request.POST.getall('meta_key')) > 0:
                        meta_keys = self.request.POST.getall('meta_key')
                        meta_values = self.request.POST.getall('meta_value')
                        for key, meta_key in enumerate(meta_keys):
                            if meta_key.strip() != '':
                                meta_value = meta_values[key].strip()
                                meta_item = ParticipantMeta(participant_id=participant_id, meta_key=meta_key, meta_value=meta_value)
                                DBSession.add(meta_item)
                                DBSession.flush()

                self.request.session.flash(_(u"Participant created."), 'success')
                return HTTPFound(location='/office/participant/view/'+str(participant.id)+'/')
            else:
                self.request.session.flash(_(u"Missing firstname and / or lastname."), 'error')
        
        
        clubs = organizationhelpers.getClubs()
        subunits = organizationhelpers.getSubUnits()
        self.request.bread.append({'url':'/office/', 'text':_('Office')})
        self.request.bread.append({'url':'/office/participant/new/', 'text':_('Create participant')})
        return {'participant':participant, 'clubs':clubs, 'subunits':subunits}
    
    @view_config(route_name='office_participant_edit', renderer='office/participant/edit.mak', permission='office_participant_edit')
    def office_participant_edit(self):
        if self.request.redirect_forbidden:
            return HTTPFound(location='/forbidden/')
        elif userhelpers.checkUserPasswordChangeNeed(self.request):
            return HTTPFound(location='/settings/me/edit/')
        _ = self.request.translate
        
        error = False
        
        participant_id = self.request.matchdict['participant_id']
        
        participant = DBSession.query(Participant).get(participant_id)
        participant.getParticipantMedicalData()
        participant.getParticipantAddressData()
        participant.getParticipantPolkuData()
        participant.getParticipantPhoneData()
        participant.getParticipantLanguageData()
        participant.getParticipantPresenceData()
        participant.getParticipantNextOfKinData()
        participant.getParticipantMetaData()

        if len(participant.address_data) == 0:
            participant.address_data.append(ParticipantAddress())
        if participant.medical_data == None:
            participant.medical_data = ParticipantMedical()
        if len(participant.next_of_kin_data) == 0:
            participant.next_of_kin_data.append(ParticipantNextOfKin())
        
        if self.request.method == 'POST':

            participant.firstname = self.request.POST.get('firstname').strip()
            participant.lastname = self.request.POST.get('lastname').strip()
            participant.nickname = self.request.POST.get('nickname').strip()
            participant.birthdate = helpers.parseFinnishDateFromString(self.request.POST.get('birthdate').strip(), False)
            participant.title = self.request.POST.get('title').strip()
            participant.age_group = self.request.POST.get('age_group').strip()
            member_no = self.request.POST.get('member_no').strip()
            if member_no == '':
                member_no = 0
            elif not member_no.isdigit():
                member_no = 0
            participant.member_no = member_no
            participant.sex = self.request.POST.get('sex').strip()
            participant.spiritual = self.request.POST.get('spiritual').strip()
            participant.club_id = self.request.POST.get('club_id').strip()
            participant.email = self.request.POST.get('email').strip()
            participant.active = 1

            subunit_id = self.request.POST.get('subunit_id').strip()
            if subunit_id > 0:
                participant.subunit_id = subunit_id
                subunit = organizationhelpers.getSubUnit(subunit_id)
                if subunit != None:
                    participant.village_id = subunit.village_id
                    village = organizationhelpers.getVillage(subunit.village_id)
                    if village != None:
                        participant.subcamp_id = village.subcamp_id
            
            DBSession.add(participant)
            DBSession.flush()
            userAudit = UserAudit(self.request.user.id)
            userAudit.model = 'Participant'
            userAudit.model_id = participant.id
            userAudit.action = 'Update'
            userAudit.revision = participant.metadata_revision
            DBSession.add(userAudit)
            DBSession.flush()
            participant_id = participant.id

            participant.address_data[0].street = self.request.POST.get('address_street').strip()
            participant.address_data[0].postalcode = self.request.POST.get('address_postalcode').strip()
            participant.address_data[0].city = self.request.POST.get('address_city').strip()
            participant.address_data[0].country = self.request.POST.get('address_country').strip()
            
            participant.next_of_kin_data[0].primary_name = self.request.POST.get('next_of_kin_0_name').strip()
            participant.next_of_kin_data[0].primary_phone = self.request.POST.get('next_of_kin_0_phone').strip()
            participant.next_of_kin_data[0].primary_email = self.request.POST.get('next_of_kin_0_email').strip()
            participant.next_of_kin_data[0].secondary_name = self.request.POST.get('next_of_kin_1_name').strip()
            participant.next_of_kin_data[0].secondary_phone = self.request.POST.get('next_of_kin_1_phone').strip()
            participant.next_of_kin_data[0].secondary_email = self.request.POST.get('next_of_kin_1_email').strip()
            
            if int(participant_id) > 0:
                participant.address_data[0].participant_id = participant_id
                DBSession.add(participant.address_data[0])
                DBSession.flush()
            
            if security.has_permission('office_participant_view_medical', self.request.context, self.request):
                if len(self.request.POST.getall('medical_diet'))>0:
                    participant.medical_data.diets = DBSession.query(ParticipantMedicalDiet).filter(ParticipantMedicalDiet.id.in_(self.request.POST.getall('medical_diet'))).all()
                                
                if len(self.request.POST.getall('medical_food_allergy'))>0:
                    participant.medical_data.food_allergies = DBSession.query(ParticipantMedicalFoodAllergy).filter(ParticipantMedicalFoodAllergy.id.in_(self.request.POST.getall('medical_food_allergy'))).all()

                participant.medical_data.additional_food = self.request.POST.get('medical_additional_food').strip()
                participant.medical_data.drugs_help = int(self.request.POST.get('medical_drugs_help'))
                participant.medical_data.illnesses = self.request.POST.get('medical_illnesses').strip()

                if len(self.request.POST.getall('medical_allergy'))>0:
                    participant.medical_data.allergies = DBSession.query(ParticipantMedicalAllergy).filter(ParticipantMedicalAllergy.id.in_(self.request.POST.getall('medical_allergy'))).all()

                participant.medical_data.additional_health = self.request.POST.get('medical_additional_health').strip()
                participant.medical_data.week_of_pregnancy = self.request.POST.get('medical_week_of_pregnancy').strip()

                if participant_id > 0:
                    participant.medical_data.participant_id = participant_id
                    DBSession.add(participant.medical_data)
                    DBSession.flush()
            else:
                participant.medical_data.participant_id = participant_id
                DBSession.add(participant.medical_data)
                DBSession.flush()

            if int(participant_id) > 0:
                language_ids = self.request.POST.getall('language_id')
                if len(participant.language_data) > 0:
                    for language in participant.language_data:
                        if str(language.id) not in language_ids:
                            DBSession.delete(language)

                if len(self.request.POST.getall('language')) > 0:
                    languages = self.request.POST.getall('language')
                    for key, lang in enumerate(languages):
                        if lang.strip() != '':
                            if int(language_ids[key]) == 0:
                                language = ParticipantLanguage(participant_id=participant_id, language=lang)
                            else:
                                language = DBSession.query(ParticipantLanguage).get(language_ids[key])
                                language.language = lang
                            DBSession.add(language)
                            DBSession.flush()

                if len(self.request.POST.getall('phone_number')) > 0:
                    phone_ids = self.request.POST.getall('phone_id')
                    if len(participant.phone_data) > 0:
                        for phone in participant.phone_data:
                            if str(phone.id) not in phone_ids:
                                DBSession.delete(phone)

                    phones = self.request.POST.getall('phone_number')
                    types = self.request.POST.getall('phone_type')
                    for key, phone_num in enumerate(phones):
                        if phone_num.strip() != '':
                            description = types[key]
                            if int(phone_ids[key]) == 0:
                                phone = ParticipantPhone(participant_id=participant_id, phone=phone_num, description=description)
                                DBSession.add(phone)
                                DBSession.flush()
                            else:
                                phone = DBSession.query(ParticipantPhone).get(phone_ids[key])
                                if phone != None:
                                    phone.phone = phone_num
                                    phone.description = description
                                    DBSession.add(phone)
                                    DBSession.flush()

                elif len(participant.phone_data) > 0:
                    for phone in participant.phone_data:
                        DBSession.delete(phone)
                    
                if len(self.request.POST.getall('meta_key')) > 0:
                    meta_ids = self.request.POST.getall('metadata_id')
                    if len(participant.meta_data) > 0:
                        for meta_data in participant.meta_data:
                            if str(meta_data.id) not in meta_ids:
                                DBSession.delete(meta_data)

                    meta_keys = self.request.POST.getall('meta_key')
                    meta_values = self.request.POST.getall('meta_value')
                    for key, meta_key in enumerate(meta_keys):
                        if meta_key.strip() != '':
                            meta_value = meta_values[key]
                            if int(meta_ids[key]) == 0:
                                meta_item = ParticipantMeta(participant_id=participant_id, meta_key=meta_key, meta_value=meta_value)
                            else:
                                
                                meta_item = DBSession.query(ParticipantMeta).get(meta_ids[key])
                                meta_item.meta_key=meta_key
                                meta_item.meta_value=meta_value
                            
                            DBSession.add(meta_item)
                            DBSession.flush()

                elif len(participant.meta_data) > 0:
                    for meta_data in participant.meta_data:
                        DBSession.delete(meta_data)

                if len(self.request.POST.getall('presence_starts')) > 0:
                    presence_ids = self.request.POST.getall('presence_id')
                    if len(participant.presence_data) > 0:
                        for presence in participant.presence_data:
                            if str(presence.id) not in presence_ids:
                                DBSession.delete(presence)
                    
                    presence_starts = self.request.POST.getall('presence_starts')
                    presence_ends = self.request.POST.getall('presence_ends')
                    presence_titles = self.request.POST.getall('presence_title')
                    presence_descriptions = self.request.POST.getall('presence_description')
                
                    for key, presence_start in enumerate(presence_starts):
                        if presence_start != '':
                            presence_starts_dt = helpers.parseFinnishDateFromString(presence_start)
                            if presence_starts_dt == False:
                                error = True
                                self.request.session.flash(_(u"Wrong time format in presences."), 'error')
                                continue
                            presence_end = presence_ends[key]

                            presence_ends_dt = None
                            if presence_end != '':
                                presence_ends_dt = helpers.parseFinnishDateFromString(presence_end)
                            
                            if presence_ends_dt == False:
                                error = True
                                self.request.session.flash(_(u"Wrong time format in presences."), 'error')
                                continue

                            presence_title = presence_titles[key]
                            presence_description = presence_descriptions[key]
                            if int(presence_ids[key]) == 0 or presence_ids[key] == '0':
                                presence = ParticipantPresence(participant_id=participant_id, starts = presence_starts_dt, ends = presence_ends_dt, title = presence_title, description = presence_description)
                                DBSession.add(presence)
                                DBSession.flush()
                            else:
                                presence = DBSession.query(ParticipantPresence).get(int(presence_ids[key]))
                                presence.presence_starts = presence_starts_dt
                                presence.presence_ends = presence_ends_dt
                                presence.title = presence_title
                                presence.description = presence_description
                                DBSession.add(presence)
                                DBSession.flush()


                elif len(participant.presence_data) > 0:
                    for presence in participant.presence_data:
                        DBSession.delete(presence)
                
            self.request.session.flash(_(u"Participant saved."), 'success')
            return HTTPFound(location='/office/participant/view/'+str(participant.id)+'/')
        
        clubs = organizationhelpers.getClubs()
        subunits = organizationhelpers.getSubUnits()
        self.request.bread.append({'url':'/office/', 'text':_('Office')})
        self.request.bread.append({'url':'/office/participant/view/'+str(participant_id)+'/', 'text':_('Participant') + ' ' + helpers.decodeString(participant.firstname) + ' ' + helpers.decodeString(participant.lastname)})
        self.request.bread.append({'url':'/office/participant/edit/'+str(participant_id)+'/', 'text':_('Edit')})
        return {'participant':participant, 'clubs':clubs, 'subunits':subunits}

    @view_config(route_name='office_participant_view', renderer='office/participant/view.mak', permission='office_participant_view')
    def office_participant_view(self):
        if self.request.redirect_forbidden:
            return HTTPFound(location='/forbidden/')
        elif userhelpers.checkUserPasswordChangeNeed(self.request):
            return HTTPFound(location='/settings/me/edit/')
        _ = self.request.translate
        
        participant_id = self.request.matchdict['participant_id']
        
        participant = DBSession.query(Participant).get(participant_id)
        
        if participant == None:
            self.request.session.flash(_(u"Participant with id ${id} not found.", mapping={'id':participant_id}), 'success')
            return HTTPFound(location='/office/')
        
        if self.request.method == 'POST' and self.request.POST.get('participant_new_status_id') != None and self.request.POST.get('participant_new_status_id').isdigit():
            participantStatus = ParticipantStatus()
            participantStatus.participant_id = participant.id
            participantStatus.status_id = int(self.request.POST.get('participant_new_status_id'))
            description = self.request.POST.get('participant_new_status_description')
            if description != None and description.strip() != '':
                participantStatus.description = description
            expected_next_change = self.request.POST.get('participant_new_status_expected_next_change')
            if expected_next_change != None and expected_next_change.strip() != '':
                participantStatus.expected_next_change = helpers.parseFinnishDateFromString(expected_next_change.strip())
            DBSession.add(participantStatus)
            DBSession.flush()
            participant.latest_status_key = int(self.request.POST.get('participant_new_status_id'))
            DBSession.add(participant)
            DBSession.flush()
            self.request.session.flash(_(u"Added new status for participant."), 'success')
            return HTTPFound(location='/office/participant/view/'+str(participant.id)+'/')
        
        
        participant.getParticipantMedicalData()
        participant.getParticipantAddressData()
        participant.getParticipantPolkuData()
        participant.getParticipantPhoneData()
        participant.getParticipantLanguageData()
        participant.getParticipantPresenceData()
        participant.getParticipantNextOfKinData()
        participant.getParticipantMetaData()

        self.request.bread.append({'url':'/office/', 'text':_('Office')})
        self.request.bread.append({'url':'/office/participant/view/'+participant_id+'/', 'text':_('Participant') + ' ' + helpers.decodeString(participant.firstname) + ' ' + helpers.decodeString(participant.lastname)})
        return {'participant':participant}

    @view_config(route_name='office_participant_view_payments', renderer='office/participant/view_payments.mak', permission='office_participant_view')
    def office_participant_view_payments(self):
        if self.request.redirect_forbidden:
            return HTTPFound(location='/forbidden/')
        elif userhelpers.checkUserPasswordChangeNeed(self.request):
            return HTTPFound(location='/settings/me/edit/')
        _ = self.request.translate

        participant_id = self.request.matchdict['participant_id']

        participant = DBSession.query(Participant).get(participant_id)
        participant.getParticipantPaymentData()
        participant.getParticipantAddressData()

        if self.request.method == 'POST':
            payment_id = self.request.POST.get('payment_id')
            is_new = False
            if payment_id != None and payment_id.strip() != '' and payment_id.strip() != '0' and payment_id.isdigit():
                payment = DBSession.query(ParticipantPayment).get(payment_id)
            else:
                payment = ParticipantPayment()
                payment.participant_id = int(participant_id)
                is_new = True
            payment.title = self.request.POST.get('payment_title').strip()
            euros = str(self.request.POST.get('payment_euros')).strip()
            euros = euros.replace(',','.')
            try:
                euros_tmp = float(euros)
            except ValueError:
                self.request.session.flash(_(u"Error converting euros-field, please check values."), 'error')
                return HTTPFound(location='/office/participant/view/'+str(participant.id)+'/payments/')
            payment.euros = float(euros)
            payment.note = self.request.POST.get('payment_note').strip()
            payment.paid = int(self.request.POST.get('payment_paid').strip())
            payment.send_invoice = int(self.request.POST.get('payment_send_invoice').strip())
            DBSession.add(payment)
            DBSession.flush()
            if is_new:
                self.request.session.flash(_(u"Added new payment for participant."), 'success')
            else:
                self.request.session.flash(_(u"Update payment for participant."), 'success')
            return HTTPFound(location='/office/participant/view/'+str(participant.id)+'/payments/')



        self.request.bread.append({'url':'/office/', 'text':_('Office')})
        self.request.bread.append({'url':'/office/participant/view/'+participant_id+'/', 'text':_('Participant') + ' ' + helpers.decodeString(participant.firstname) + ' ' + helpers.decodeString(participant.lastname)})
        self.request.bread.append({'url':'/office/participant/view/'+participant_id+'/payments/', 'text':_('Payments')})
        return {'participant':participant}
    
    @view_config(route_name='office_participant_cancel', renderer='office/participant/view.mak', permission='office_participant_edit')
    def office_participant_cancel(self):
        if self.request.redirect_forbidden:
            return HTTPFound(location='/forbidden/')
        elif userhelpers.checkUserPasswordChangeNeed(self.request):
            return HTTPFound(location='/settings/me/edit/')
        _ = self.request.translate
        
        participant_id = self.request.matchdict['participant_id']
        participant = DBSession.query(Participant).get(participant_id)
        participant.active = False
        DBSession.add(participant)
        DBSession.flush()
        self.request.session.flash(_(u"Canceled participant booking."), 'success')
        return HTTPFound(location='/office/participant/view/'+str(participant.id)+'/')
        
        
    @view_config(route_name='office_participant_uncancel', renderer='office/participant/view.mak', permission='office_participant_edit')
    def office_participant_uncancel(self):
        if self.request.redirect_forbidden:
            return HTTPFound(location='/forbidden/')
        elif userhelpers.checkUserPasswordChangeNeed(self.request):
            return HTTPFound(location='/settings/me/edit/')
        _ = self.request.translate

        participant_id = self.request.matchdict['participant_id']
        participant = DBSession.query(Participant).get(participant_id)
        participant.active = True
        DBSession.add(participant)
        DBSession.flush()
        self.request.session.flash(_(u"Uncanceled participant booking."), 'success')
        return HTTPFound(location='/office/participant/view/'+str(participant.id)+'/')
        
    
    @view_config(route_name='office_participant_search', renderer='office/participant/search.mak', permission='office_participant_view')
    def office_participant_search(self):
        if self.request.redirect_forbidden:
            return HTTPFound(location='/forbidden/')
        elif userhelpers.checkUserPasswordChangeNeed(self.request):
            return HTTPFound(location='/settings/me/edit/')
        _ = self.request.translate
        
        participants = []
        search_string = ''

        if self.request.method == 'POST':
            search_string = self.request.POST.get('searchstr').strip()

            if search_string == '':
                self.request.session.flash(_(u"Empty search, please provide search string."), 'error')
            else:
                participants = DBSession.query(Participant).filter(
                    or_(
                    Participant.firstname.like('%'+search_string+'%'),
                    Participant.lastname.like('%'+search_string+'%'),
                    Participant.nickname.like('%'+search_string+'%'),
                    Participant.member_no.like('%'+search_string+'%'),
                    Participant.booking_no.like('%'+search_string+'%'),
                    )).all()

        self.request.bread.append({'url':'/office/', 'text':_('Office')})
        self.request.bread.append({'url':'/office/participant/search/', 'text':_('Search')})
        return {'participants':participants, 'search_string':search_string}

    @view_config(route_name='office_report', renderer='office/report.mak', permission='office_view')
    def office_report(self):
        if self.request.redirect_forbidden:
            return HTTPFound(location='/forbidden/')
        elif userhelpers.checkUserPasswordChangeNeed(self.request):
            return HTTPFound(location='/settings/me/edit/')
        _ = self.request.translate
        

        # if self.request.matched_route.name == 'office_report_excel':
        #     self.request.response.content_type = "application/ms-excel"
        #     self.request.response.content_disposition = 'attachment; filename="leirirekkari_report_'+datetime.datetime.now().strftime("%Y%m%d_%H%M")+'.xls"'

        if self.request.POST.get('mass_tools_submitter') != None and self.request.POST.get('mass_tools_action') != None and self.request.POST.get('mass_tools_action') != '':
            participant_ids_to_update = self.request.POST.getall('participant_id_checkbox')
            action = self.request.POST.get('mass_tools_action')
            if action == 'status':
                for participant_id in participant_ids_to_update:
                    participant = DBSession.query(Participant).get(int(participant_id))
                    if participant.id == int(participant_id):
                        participantStatus = ParticipantStatus()
                        participantStatus.participant_id = participant.id
                        participantStatus.status_id = int(self.request.POST.get('participant_new_status_id'))
                        description = self.request.POST.get('participant_new_status_description')
                        if description != None and description.strip() != '':
                            participantStatus.description = description
                        expected_next_change = self.request.POST.get('participant_new_status_expected_next_change')
                        if expected_next_change != None and expected_next_change.strip() != '':
                            participantStatus.expected_next_change = helpers.parseFinnishDateFromString(expected_next_change.strip())
                        DBSession.add(participantStatus)
                        self.request.session.flash(_(u"Added new status for participant ${name}.", mapping={'name':helpers.decodeString(participant.firstname) + ' ' + helpers.decodeString(participant.lastname)}), 'success')
        
            DBSession.flush()
            
        ret = self.getReportResults(self.request)
        
        participants = ret['participants']
        filter_strings = ret['filter_strings']
        to_view = ret['to_view']
        

        self.request.bread.append({'url':'/office/', 'text':_('Office')})
        self.request.bread.append({'url':'/office/report/', 'text':_('Report')})
        return {'participants':participants, 'filter_strings':filter_strings, 'to_view':to_view}
    
    def setExcelHeaderCell(self, ws, col_counter, row_counter, value, width=None):
        col = get_column_letter(col_counter)
        
        cell = ws.cell('%s%s'%(col, row_counter))
        cell.value = helpers.decodeString(value)
        cell.style.font.name = 'Arial'
        cell.style.font.size = 11
        cell.style.font.bold = True
        cell.style.alignment.wrap_text = True
        if width != None:
            ws.column_dimensions[col].width = width
        else:
            ws.column_dimensions[col].width = 16
        
        
    def setExcelCell(self, ws, col_counter, row_counter, value, type=None):
        col = get_column_letter(col_counter)

        cell = ws.cell('%s%s'%(col, row_counter))
        if type != None:
            cell.set_value_explicit(value, type)
        else:
            cell.value = helpers.decodeString(value)
        cell.style.font.name = 'Arial'
        cell.style.font.size = 11
        cell.style.font.bold = False
        cell.style.alignment.wrap_text = True

    def setExcelCellFill(self, ws, col_counter, row_counter, color):
        col = get_column_letter(col_counter)

        cell = ws.cell('%s%s'%(col, row_counter))
        cell.style.fill.fill_type = openpyxl_style.Fill.FILL_SOLID
        cell.style.fill.start_color.index = color
        
    
    @view_config(route_name='office_report_excel', renderer='office/report_excel.mak', permission='office_view')
    def office_report_excel(self):
        if self.request.redirect_forbidden:
            return HTTPFound(location='/forbidden/')
        elif userhelpers.checkUserPasswordChangeNeed(self.request):
            return HTTPFound(location='/settings/me/edit/')
        _ = self.request.translate
        
        base_path_obj = DBSession.query(Setting).filter(Setting.setting_key == 'report_file_creation_dir').first()
        base_path = base_path_obj.setting_value
        
        dest_filename = u'leirirekkari_report_'+datetime.datetime.now().strftime("%Y%m%d_%H%M")+'_'+str(self.request.user.id)+'.xlsx'

        self.request.response.content_disposition = 'attachment; filename="'+dest_filename+'"'
        
        ret = self.getReportResults(self.request)
        
        participants = ret['participants']
        filter_strings = ret['filter_strings']
        to_view = ret['to_view']
        
        subcamps = organizationhelpers.getSubcamps()
        villages = organizationhelpers.getVillages()
        village_kitchens = organizationhelpers.getVillageKitchens()
        subunits = organizationhelpers.getSubUnits()
        clubs = organizationhelpers.getClubs()

        stats = {
            'agegroups':{
                0:0,
                1:0,
                2:0,
                3:0,
                4:0,
                5:0,
                6:0,
                7:0,
            },
            'spiritual':{
                0:0,
                10:0,
                20:0,
            },
            'sex':{
                0:0,
                10:0,
                20:0,
            },
            'total':0
        }

        kitchen_stats = {
            'allergies': {},
            'food_allergies': {},
            'diets': {},
        }
        
        kitchen_table = {
            'allergies': {},
            'food_allergies': {},
            'diets': {},
            'participants': {}
        }
        
        
        
        wb = Workbook()
        wb.properties.creator = self.request.user.firstname + ' ' + self.request.user.lastname
        wb.properties.title = r'Saraste 2012 leirirekkari raportti'
        wb.properties.subject = r'Saraste 2012 leirirekkari raportti'
        wb.properties.company = _('Papa name')
        
        
        ws = wb.worksheets[0]
        sheet_index = 0
        
        ws.title = "Raportti"
        
        col_counter = 1
        row_counter = 1
        
        
        col = get_column_letter(col_counter)

        cell = ws.cell('%s%s'%(col, row_counter))
        cell.value = _(u"Leirirekkari raportti")
        cell.style.font.name = 'Arial'
        cell.style.font.size = 14
        cell.style.font.bold = True
        cell.style.alignment.wrap_text = False

        row_counter += 1
        row_counter += 1
        
        cell = ws.cell('%s%s'%(col, row_counter))
        cell.value = helpers.modDateTime(datetime.datetime.now())
        cell.style.font.name = 'Arial'
        cell.style.font.size = 11
        cell.style.font.bold = True
        cell.style.alignment.wrap_text = False
        
        row_counter += 1
        row_counter += 1
        
        
        if 'show_status' in filter_strings and filter_strings['show_status'] != 'No':
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Is canceled'))
            col_counter += 1
            
        if 'booking_no' in to_view and to_view['booking_no']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Polku booking no'))
            col_counter += 1

        if 'polku_member_id' in to_view and to_view['polku_member_id']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Polku member id'))
            col_counter += 1

        if 'basic_info' in to_view and to_view['basic_info']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Firstname'), 20)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Lastname'), 20)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Nickname'), 20)
            col_counter += 1

        if 'designation_title' in to_view and to_view['designation_title']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Designation'), 20)
            col_counter += 1
        
        if 'address' in to_view and to_view['address']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Street'), 25)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Postalcode'), 10)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('City'), 10)
            col_counter += 1

        if 'phone_and_email' in to_view and to_view['phone_and_email']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Phone'), 20)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Email'), 20)
            col_counter += 1

        if 'languages' in to_view and to_view['languages']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Languages'), 10)
            col_counter += 1

        if 'status' in to_view and to_view['status']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Status'), 15)
            col_counter += 1

        if 'status_all' in to_view and to_view['status_all']:
            if 'status' not in to_view or not to_view['status']:
                self.setExcelHeaderCell(ws, col_counter, row_counter, _('Status'), 15)
                col_counter += 1

            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Expected next change'), 10)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Status description'), 20)
            col_counter += 1

        if 'sex' in to_view and to_view['sex']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Sex'), 10)
            col_counter += 1

        if 'age_group' in to_view and to_view['age_group']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Agegroup'), 10)
            col_counter += 1

        if 'birthdate' in to_view and to_view['birthdate']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Birthdate'), 15)
            col_counter += 1

        if 'age' in to_view and to_view['age']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Age'), 5)
            col_counter += 1

        if 'club' in to_view and to_view['club']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Club'), 20)
            col_counter += 1

        if 'subcamp' in to_view and to_view['subcamp']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Subcamp'), 20)
            col_counter += 1

        if 'village' in to_view and to_view['village']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Village'), 20)
            col_counter += 1

        if 'village_kitchen' in to_view and to_view['village_kitchen']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Village kitchen'), 20)
            col_counter += 1

        if 'subunit' in to_view and to_view['subunit']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Subunit'), 20)
            col_counter += 1

        if 'presence' in to_view and to_view['presence']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Presences'), 40)
            col_counter += 1

        if 'spiritual' in to_view and to_view['spiritual']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Spiritual'), 20)
            col_counter += 1

        if ('medical_diets' in to_view and to_view['medical_diets']) or ('kitchen_table' in to_view and to_view['kitchen_table']):
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Medical_diet'), 20)
            col_counter += 1

        if 'medical_diets_boolean' in to_view and to_view['medical_diets_boolean']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Medical_diet_boolean'), 10)
            col_counter += 1

        if ('medical_food_allergies' in to_view and to_view['medical_food_allergies']) or ('kitchen_table' in to_view and to_view['kitchen_table']):
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Medical_food_allergy'), 20)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Medical_additional_food'), 20)
            col_counter += 1

        if 'medical_food_allergies_boolean' in to_view and to_view['medical_food_allergies_boolean']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Medical_food_allergy_boolean'), 10)
            col_counter += 1

        if 'medical_allergies' in to_view and to_view['medical_allergies']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Allergies'), 20)
            col_counter += 1

        if 'medical_allergies_boolean' in to_view and to_view['medical_allergies_boolean']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Allergies_boolean'), 10)
            col_counter += 1

        if 'medical_other' in to_view and to_view['medical_other']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Medical_drugs_help'), 10)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Medical_illnesses'), 20)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Medical_additional_health'), 20)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Pregnancy weeks during camp'), 5)
            col_counter += 1

        if 'medical_need_help_boolean' in to_view and to_view['medical_need_help_boolean']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _('Medical_drugs_help_boolean'), 10)
            col_counter += 1

        if 'next_of_kin' in to_view and to_view['next_of_kin']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Next of kin, primary"), 20)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Next of kin, primary phone"), 20)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Next of kin, secondary"), 20)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Next of kin, secondary phone"), 20)
            col_counter += 1

        if 'wishes' in to_view and to_view['wishes']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Wishes activity 1"), 15)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Wishes activity 2"), 15)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Wishes activity 3"), 15)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Preliminary signups"), 15)
            col_counter += 1

        if 'payments' in to_view and to_view['payments']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Payments total"), 15)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Paid total"), 15)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"To pay yet"), 15)
            col_counter += 1

        if 'designation_all' in to_view and to_view['designation_all']:
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Job at camp"), 15)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Enlisted by"), 15)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Enlister works as"), 15)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Enlistment table a"), 15)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Enlistment table b1"), 15)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Enlistment table b2"), 15)
            col_counter += 1
            self.setExcelHeaderCell(ws, col_counter, row_counter, _(u"Enlistment table b3"), 15)
            col_counter += 1

        
        total_cols = col_counter - 1
        
        if len(participants) > 0:
            for participant in participants:
                col_counter = 1
                row_counter += 1

                stats['total'] += 1
                stats['agegroups'][participant.age_group] += 1
                stats['sex'][participant.sex] += 1
                stats['spiritual'][participant.spiritual] += 1
                
                
                if 'show_status' in filter_strings and filter_strings['show_status'] != 'No':
                    if not participant.active:
                        self.setExcelCell(ws, col_counter, row_counter, _(u"Canceled"))
                    col_counter += 1

                if 'booking_no' in to_view and to_view['booking_no']:
                    if participant.booking_no != None:
                        self.setExcelCell(ws, col_counter, row_counter, participant.booking_no.strip('|').replace('|', ', '))
                    col_counter += 1

                if 'polku_member_id' in to_view and to_view['polku_member_id']:
                    self.setExcelCell(ws, col_counter, row_counter, participant.member_id)
                    col_counter += 1

                if 'basic_info' in to_view and to_view['basic_info']:
                    self.setExcelCell(ws, col_counter, row_counter, participant.firstname)
                    col_counter += 1
                    self.setExcelCell(ws, col_counter, row_counter, participant.lastname)
                    col_counter += 1
                    self.setExcelCell(ws, col_counter, row_counter, participant.nickname)
                    col_counter += 1

                if 'designation_title' in to_view and to_view['designation_title']:
                    self.setExcelCell(ws, col_counter, row_counter, participant.title)
                    col_counter += 1

                if 'address' in to_view and to_view['address']:
                    participant.getParticipantAddressData()
                    if len(participant.address_data) == 0:
                        col_counter += 1
                        col_counter += 1
                        col_counter += 1
                    else:
                        address = participant.address_data[0]
                        self.setExcelCell(ws, col_counter, row_counter, address.street)
                        col_counter += 1
                        self.setExcelCell(ws, col_counter, row_counter, address.postalcode)
                        col_counter += 1
                        self.setExcelCell(ws, col_counter, row_counter, address.city)
                        col_counter += 1
            
                if 'phone_and_email' in to_view and to_view['phone_and_email']:
                    participant.getParticipantPhoneData()
                    tmp = ''
                    if len(participant.phone_data) == 0:
                        for phone in participant.phone_data:
                            tmp += phone.phone + ', '
                    self.setExcelCell(ws, col_counter, row_counter, tmp.strip(','))
                    col_counter += 1
                    self.setExcelCell(ws, col_counter, row_counter, participant.email)
                    col_counter += 1

                if 'languages' in to_view and to_view['languages']:
                    participant.getParticipantLanguageData()
                    tmp = ''
                    if len(participant.language_data) == 0:
                        for language in participant.language_data:
                            tmp += language.language + ', '
                    self.setExcelCell(ws, col_counter, row_counter, tmp.strip(','))
                    col_counter += 1
                
                if 'status' in to_view and to_view['status']:
                    participant.getParticipantStatus()
                    if participant.status == None:
                        self.setExcelCell(ws, col_counter, row_counter, _("Participant status: 0"))
                        col_counter += 1
                    else:
                        self.setExcelCell(ws, col_counter, row_counter, _("Participant status: "+str(participant.status.status_id)))
                        col_counter += 1
                
                if 'status_all' in to_view and to_view['status_all']:
                    participant.getParticipantStatus()
                    if participant.status == None:
                        if 'status' not in to_view or not to_view['status']:
                            self.setExcelCell(ws, col_counter, row_counter, _("Participant status: 0"))
                            col_counter += 1
                        col_counter += 1
                        col_counter += 1
                    else:
                        if 'status' not in to_view or not to_view['status']:
                            self.setExcelCell(ws, col_counter, row_counter, _("Participant status: "+str(participant.status.status_id)))
                            col_counter += 1
                        
                        if participant.status.expected_next_change != None:
                            self.setExcelCell(ws, col_counter, row_counter, helpers.modDateTime(participant.status.expected_next_change, 'shortwithtime'))
                        col_counter += 1

                        self.setExcelCell(ws, col_counter, row_counter, helpers.literal(participant.status.description))
                        col_counter += 1
    
                if 'sex' in to_view and to_view['sex']:
                    if participant.sex == 10:
                        tmp = _(u"Men")
                    elif participant.sex == 20:
                        tmp = _(u"Female")
                    else:
                        tmp = _(u"Unknown")
                    self.setExcelCell(ws, col_counter, row_counter, tmp)
                    col_counter += 1
                
                if 'age_group' in to_view and to_view['age_group']:
                    self.setExcelCell(ws, col_counter, row_counter, _(u"age_group_"+str(participant.age_group)))
                    col_counter += 1
                
                if 'birthdate' in to_view and to_view['birthdate']:
                    self.setExcelCell(ws, col_counter, row_counter, helpers.modDateTime(participant.birthdate, 'short'))
                    col_counter += 1
                
                if 'age' in to_view and to_view['age']:
                    self.setExcelCell(ws, col_counter, row_counter, helpers.calculateAgeInYears(participant.birthdate))
                    col_counter += 1

                if 'club' in to_view and to_view['club']:
                    self.setExcelCell(ws, col_counter, row_counter, organizationhelpers.getClubName(participant.club_id))
                    col_counter += 1
                
                if 'subcamp' in to_view and to_view['subcamp']:
                    self.setExcelCell(ws, col_counter, row_counter, organizationhelpers.getSubcampName(participant.subcamp_id))
                    col_counter += 1
                
                if 'village' in to_view and to_view['village']:
                    self.setExcelCell(ws, col_counter, row_counter, organizationhelpers.getVillageName(participant.village_id))
                    col_counter += 1
                
                if 'village_kitchen' in to_view and to_view['village_kitchen']:
                    self.setExcelCell(ws, col_counter, row_counter, organizationhelpers.getVillageKitchenNameByVillage(participant.village_id))
                    col_counter += 1
                
                if 'subunit' in to_view and to_view['subunit']:
                    self.setExcelCell(ws, col_counter, row_counter, organizationhelpers.getSubUnitName(participant.subunit_id))
                    col_counter += 1
                
                if 'presence' in to_view and to_view['presence']:
                    participant.getParticipantPresenceData()
                    tmp = ''
                    if len(participant.presence_data)>0:
                        for presence in participant.presence_data:
                            tmp += helpers.getDayString(presence.presence_starts, 'short') + ' ' + helpers.modDateTime(presence.presence_starts, 'shortwithtime') + ' - ' + helpers.getDayString(presence.presence_ends, 'short') + ' ' + helpers.modDateTime(presence.presence_ends, 'shortwithtime') + "\n"
                    self.setExcelCell(ws, col_counter, row_counter, tmp.strip("\n"))
                    col_counter += 1

                if 'spiritual' in to_view and to_view['spiritual']:
                    if participant.spiritual == 10:
                        self.setExcelCell(ws, col_counter, row_counter, _(u"The ecumenical church service"))
                    elif participant.spiritual == 20:
                        self.setExcelCell(ws, col_counter, row_counter, _(u"Life stance programme"))
                    col_counter += 1
                
                if ('medical_diets' in to_view and to_view['medical_diets']) or ('kitchen_table' in to_view and to_view['kitchen_table']):
                    if not participant.medical_data_searched:
                        participant.getParticipantMedicalData()

                    if participant.medical_data != None:
                        if len(participant.medical_data.diets) > 0:
                            medical_data = participant.medical_data
                            tmp = ''
                            for diet in medical_data.diets:
                                if diet.name not in kitchen_table['diets']:
                                    kitchen_table['diets'][diet.name] = []
                                kitchen_table['diets'][diet.name].append(participant.id)
                                if participant.id not in kitchen_table['participants']:
                                    kitchen_table['participants'][participant.id] = participant.firstname + ' ' + participant.lastname
                                
                                if diet.name in kitchen_stats['diets']:
                                    kitchen_stats['diets'][diet.name] += 1
                                else:
                                    kitchen_stats['diets'][diet.name] = 1
                                tmp += diet.name + "\n"
                            self.setExcelCell(ws, col_counter, row_counter, tmp.strip("\n"))
                    col_counter += 1
                
                if 'medical_diets_boolean' in to_view and to_view['medical_diets_boolean']:
                    if not participant.medical_data_searched:
                        participant.getParticipantMedicalData()

                    if participant.medical_data != None:
                        if len(participant.medical_data.diets) > 0:
                            self.setExcelCell(ws, col_counter, row_counter, _(u"Yes"))
                        else:
                            self.setExcelCell(ws, col_counter, row_counter, _(u"No"))
                    col_counter += 1
                
                if ('medical_food_allergies' in to_view and to_view['medical_food_allergies']) or ('kitchen_table' in to_view and to_view['kitchen_table']):
                    if not participant.medical_data_searched:
                        participant.getParticipantMedicalData()

                    if participant.medical_data != None:
                        if len(participant.medical_data.food_allergies) > 0:
                            medical_data = participant.medical_data
                            tmp = ''
                            for food_allergy in medical_data.food_allergies:
                                if food_allergy.name not in kitchen_table['food_allergies']:
                                    kitchen_table['food_allergies'][food_allergy.name] = []
                                kitchen_table['food_allergies'][food_allergy.name].append(participant.id)
                                if participant.id not in kitchen_table['participants']:
                                    kitchen_table['participants'][participant.id] = participant.firstname + ' ' + participant.lastname
                                
                                if food_allergy.name in kitchen_stats['food_allergies']:
                                    kitchen_stats['food_allergies'][food_allergy.name] += 1
                                else:
                                    kitchen_stats['food_allergies'][food_allergy.name] = 1
                                tmp += food_allergy.name + "\n"
                            self.setExcelCell(ws, col_counter, row_counter, tmp.strip("\n"))
                    col_counter += 1
                    if participant.medical_data != None:
                        self.setExcelCell(ws, col_counter, row_counter, participant.medical_data.additional_food)
                    col_counter += 1
                
                
                if 'medical_food_allergies_boolean' in to_view and to_view['medical_food_allergies_boolean']:
                    if not participant.medical_data_searched:
                        participant.getParticipantMedicalData()

                    if participant.medical_data != None:
                        if len(participant.medical_data.food_allergies) > 0:
                            self.setExcelCell(ws, col_counter, row_counter, _(u"Yes"))
                        else:
                            self.setExcelCell(ws, col_counter, row_counter, _(u"No"))
                    col_counter += 1
                
                
                if 'medical_allergies' in to_view and to_view['medical_allergies']:
                    if not participant.medical_data_searched:
                        participant.getParticipantMedicalData()

                    if participant.medical_data != None:
                        if len(participant.medical_data.allergies) > 0:
                            medical_data = participant.medical_data
                            tmp = ''
                            for allergy in medical_data.allergies:
                                if allergy.name in kitchen_stats['allergies']:
                                    kitchen_stats['allergies'][allergy.name] += 1
                                else:
                                    kitchen_stats['allergies'][allergy.name] = 1
                                tmp += allergy.name + "\n"
                            self.setExcelCell(ws, col_counter, row_counter, tmp.strip("\n"))
                    col_counter += 1
                
                if 'medical_allergies_boolean' in to_view and to_view['medical_allergies_boolean']:
                    if not participant.medical_data_searched:
                        participant.getParticipantMedicalData()

                    if participant.medical_data != None:
                        if len(participant.medical_data.allergies) > 0:
                            self.setExcelCell(ws, col_counter, row_counter, _(u"Yes"))
                        else:
                            self.setExcelCell(ws, col_counter, row_counter, _(u"No"))
                    col_counter += 1
                
                
                if 'medical_other' in to_view and to_view['medical_other']:
                    if not participant.medical_data_searched:
                        participant.getParticipantMedicalData()
                    if participant.medical_data == None:
                        col_counter += 1
                        col_counter += 1
                        col_counter += 1
                        col_counter += 1
                    else:
                        if participant.medical_data.drugs_help == 10:
                            self.setExcelCell(ws, col_counter, row_counter, _("Own leader or subunit first aid"))
                        elif participant.medical_data.drugs_help == 20:
                            self.setExcelCell(ws, col_counter, row_counter, _("Camphospital"))
                        else:
                            self.setExcelCell(ws, col_counter, row_counter, _("Dont need"))
                        col_counter += 1
                        self.setExcelCell(ws, col_counter, row_counter, participant.medical_data.illnesses)
                        col_counter += 1
                        self.setExcelCell(ws, col_counter, row_counter, participant.medical_data.additional_health)
                        col_counter += 1
                        if participant.medical_data.week_of_pregnancy != '':
                            self.setExcelCell(ws, col_counter, row_counter, participant.medical_data.week_of_pregnancy)
                        col_counter += 1
                
                if 'medical_need_help_boolean' in to_view and to_view['medical_need_help_boolean']:
                    if not participant.medical_data_searched:
                        participant.getParticipantMedicalData()
                    if participant.medical_data != None:
                        if participant.medical_data.drugs_help == 10:
                            self.setExcelCell(ws, col_counter, row_counter, _("Own leader or subunit first aid"))
                        elif participant.medical_data.drugs_help == 20:
                            self.setExcelCell(ws, col_counter, row_counter, _("Camphospital"))
                        else:
                            self.setExcelCell(ws, col_counter, row_counter, _("Dont need"))
                    col_counter += 1
                
                
                if 'next_of_kin' in to_view and to_view['next_of_kin']:
                    participant.getParticipantNextOfKinData()
                    if len(participant.next_of_kin_data) == 0:
                        col_counter += 1
                        col_counter += 1
                        col_counter += 1
                        col_counter += 1
                    else:
                        tmp = helpers.checkString(participant.next_of_kin_data[0].primary_name) + "\n" + helpers.checkString(participant.next_of_kin_data[0].primary_phone) + "\n" + helpers.checkString(participant.next_of_kin_data[0].primary_email)
                        self.setExcelCell(ws, col_counter, row_counter, tmp)
                        col_counter += 1
                        self.setExcelCell(ws, col_counter, row_counter, participant.next_of_kin_data[0].primary_phone)
                        col_counter += 1
                        tmp = helpers.checkString(participant.next_of_kin_data[0].secondary_name) + "\n" + helpers.checkString(participant.next_of_kin_data[0].secondary_phone) + "\n" + helpers.checkString(participant.next_of_kin_data[0].secondary_email)
                        self.setExcelCell(ws, col_counter, row_counter, tmp)
                        col_counter += 1
                        self.setExcelCell(ws, col_counter, row_counter, participant.next_of_kin_data[0].secondary_phone)
                        col_counter += 1
                
                if 'wishes' in to_view and to_view['wishes']:
                    participant.getParticipantWishes()

                    if participant.wishes == None:
                        col_counter += 1
                        col_counter += 1
                        col_counter += 1
                        col_counter += 1
                    else:
                        self.setExcelCell(ws, col_counter, row_counter, participant.wishes.activity_1.name)
                        col_counter += 1
                        self.setExcelCell(ws, col_counter, row_counter, participant.wishes.activity_2.name)
                        col_counter += 1
                        self.setExcelCell(ws, col_counter, row_counter, participant.wishes.activity_3.name)
                        col_counter += 1
                        if len(participant.wishes.preliminary_signups) > 0:
                            tmp = ''
                            for signup in participant.wishes.preliminary_signups:
                                tmp += signup.name + "\n"
                            self.setExcelCell(ws, col_counter, row_counter, tmp.strip("\n"))
                        col_counter += 1

                if 'payments' in to_view and to_view['payments']:
                    participant.getParticipantPaymentData()
                    paymentSums = participanthelpers.countPaymentSums(participant)

                    self.setExcelCell(ws, col_counter, row_counter, paymentSums['payments_total'])
                    col_counter += 1
                    self.setExcelCell(ws, col_counter, row_counter, paymentSums['paid_total'])
                    col_counter += 1
                    self.setExcelCell(ws, col_counter, row_counter, paymentSums['to_pay_total'])
                    col_counter += 1
                
                if 'designation_all' in to_view and to_view['designation_all']:
                    participant.getParticipantEnlistment()
                
                    if participant.enlistment == None:
                        col_counter += 1
                        col_counter += 1
                        col_counter += 1
                        col_counter += 1
                        col_counter += 1
                        col_counter += 1
                        col_counter += 1
                    else:
                        self.setExcelCell(ws, col_counter, row_counter, participant.enlistment.job_at_camp)
                        col_counter += 1
                        self.setExcelCell(ws, col_counter, row_counter, participant.enlistment.enlisted_by)
                        col_counter += 1
                        self.setExcelCell(ws, col_counter, row_counter, participant.enlistment.enlister_works_as)
                        col_counter += 1
                        self.setExcelCell(ws, col_counter, row_counter, participanthelpers.getEnlistmentName(participant.enlistment.enlistment_table_a_id))
                        col_counter += 1
                        self.setExcelCell(ws, col_counter, row_counter, participanthelpers.getEnlistmentName(participant.enlistment.enlistment_table_b1_id))
                        col_counter += 1
                        self.setExcelCell(ws, col_counter, row_counter, participanthelpers.getEnlistmentName(participant.enlistment.enlistment_table_b2_id))
                        col_counter += 1
                        self.setExcelCell(ws, col_counter, row_counter, participanthelpers.getEnlistmentName(participant.enlistment.enlistment_table_b3_id))
                        col_counter += 1
                
        sheet_index += 1
        wb.create_sheet(sheet_index)
        ws1 = wb.worksheets[sheet_index]

        ws1.title = "Tilastot"

        col_counter = 1
        row_counter = 1


        col = get_column_letter(col_counter)

        cell = ws1.cell('%s%s'%(col, row_counter))
        cell.value = _(u"Leirirekkari raportti, tilastot")
        cell.style.font.name = 'Arial'
        cell.style.font.size = 14
        cell.style.font.bold = True
        cell.style.alignment.wrap_text = False

        row_counter += 1
        row_counter += 1

        cell = ws1.cell('%s%s'%(col, row_counter))
        cell.value = helpers.modDateTime(datetime.datetime.now())
        cell.style.font.name = 'Arial'
        cell.style.font.size = 11
        cell.style.font.bold = True
        cell.style.alignment.wrap_text = False

        row_counter += 1
        row_counter += 1
        
        self.setExcelHeaderCell(ws1, col_counter, row_counter, _(u"Total"))
        col_counter += 1
        self.setExcelCell(ws1, col_counter, row_counter, stats['total'])

        col_counter = 1
        
        row_counter += 1
        row_counter += 1
        self.setExcelHeaderCell(ws1, col_counter, row_counter, _(u"Age groups"))
        col_counter += 1

        for key, value in enumerate(stats['agegroups']):
            row_counter += 1
            col_counter = 1
            self.setExcelHeaderCell(ws1, col_counter, row_counter, _(u"age_group_"+str(value)))
            col_counter += 1
            self.setExcelCell(ws1, col_counter, row_counter, stats['agegroups'][value])

        col_counter = 1
        row_counter += 1

        row_counter += 1
        row_counter += 1
        self.setExcelHeaderCell(ws1, col_counter, row_counter, _(u"Sex"))
        col_counter += 1

        for key, value in enumerate(stats['sex']):
            row_counter += 1
            col_counter = 1
            self.setExcelHeaderCell(ws1, col_counter, row_counter, _(u"sex_"+str(value)))
            col_counter += 1
            self.setExcelCell(ws1, col_counter, row_counter, stats['sex'][value])

        col_counter = 1
        row_counter += 1

        row_counter += 1
        row_counter += 1
        self.setExcelHeaderCell(ws1, col_counter, row_counter, _(u"Spiritual"))
        col_counter += 1

        for key, value in enumerate(stats['spiritual']):
            row_counter += 1
            col_counter = 1
            self.setExcelHeaderCell(ws1, col_counter, row_counter, _(u"spiritual_"+str(value)))
            col_counter += 1
            self.setExcelCell(ws1, col_counter, row_counter, stats['spiritual'][value])

        col_counter = 1
        row_counter += 1


        if ('medical_diets' in to_view and to_view['medical_diets']) or ('medical_food_allergies' in to_view and to_view['medical_food_allergies']) or ('medical_allergies' in to_view and to_view['medical_allergies']):
            sheet_index += 1
            wb.create_sheet(sheet_index)
            ws2 = wb.worksheets[sheet_index]

            ws2.title = "Allergiatilastot"

            col_counter = 1
            row_counter = 1


            col = get_column_letter(col_counter)

            cell = ws2.cell('%s%s'%(col, row_counter))
            cell.value = _(u"Leirirekkari raportti, allergiatilastot")
            cell.style.font.name = 'Arial'
            cell.style.font.size = 14
            cell.style.font.bold = True
            cell.style.alignment.wrap_text = False

            row_counter += 1
            row_counter += 1

            cell = ws2.cell('%s%s'%(col, row_counter))
            cell.value = helpers.modDateTime(datetime.datetime.now())
            cell.style.font.name = 'Arial'
            cell.style.font.size = 11
            cell.style.font.bold = True
            cell.style.alignment.wrap_text = False

            row_counter += 1
            row_counter += 1
            
            if ('medical_diets' in to_view and to_view['medical_diets']) or ('kitchen_table' in to_view and to_view['kitchen_table']):
                col_counter = 1
                row_counter += 1

                row_counter += 1
                row_counter += 1
                self.setExcelHeaderCell(ws2, col_counter, row_counter, _(u"Medical_diet"))
                col_counter += 1

                for key in kitchen_stats['diets']:
                    row_counter += 1
                    col_counter = 1
                    self.setExcelHeaderCell(ws2, col_counter, row_counter, key)
                    col_counter += 1
                    self.setExcelCell(ws2, col_counter, row_counter, str(kitchen_stats['diets'][key]))

                col_counter = 1
                row_counter += 1

            if ('medical_food_allergies' in to_view and to_view['medical_food_allergies']) or ('kitchen_table' in to_view and to_view['kitchen_table']):
                col_counter = 1
                row_counter += 1

                row_counter += 1
                row_counter += 1
                self.setExcelHeaderCell(ws2, col_counter, row_counter, _(u"Medical_food_allergy"))
                col_counter += 1

                for key in kitchen_stats['food_allergies']:
                    row_counter += 1
                    col_counter = 1
                    self.setExcelHeaderCell(ws2, col_counter, row_counter, key)
                    col_counter += 1
                    self.setExcelCell(ws2, col_counter, row_counter, str(kitchen_stats['food_allergies'][key]))

                col_counter = 1
                row_counter += 1

            if ('medical_allergies' in to_view and to_view['medical_allergies']) or ('kitchen_table' in to_view and to_view['kitchen_table']):
                col_counter = 1
                row_counter += 1

                row_counter += 1
                row_counter += 1
                self.setExcelHeaderCell(ws2, col_counter, row_counter, _(u"Allergies"))
                col_counter += 1

                for key in kitchen_stats['allergies']:
                    row_counter += 1
                    col_counter = 1
                    self.setExcelHeaderCell(ws2, col_counter, row_counter, key)
                    col_counter += 1
                    self.setExcelCell(ws2, col_counter, row_counter, str(kitchen_stats['allergies'][key]))

                col_counter = 1
                row_counter += 1


        if 'kitchen_table' in to_view and to_view['kitchen_table']:
            
            sheet_index += 1
            wb.create_sheet(sheet_index)
            ws3 = wb.worksheets[sheet_index]

            ws3.title = "Allergiat ja ruokavaliot"

            col_counter = 1
            row_counter = 1


            col = get_column_letter(col_counter)

            cell = ws3.cell('%s%s'%(col, row_counter))
            cell.value = _(u"Leirirekkari raportti, allergiat ja ruokavaliot")
            cell.style.font.name = 'Arial'
            cell.style.font.size = 14
            cell.style.font.bold = True
            cell.style.alignment.wrap_text = False

            row_counter += 1
            row_counter += 1

            cell = ws3.cell('%s%s'%(col, row_counter))
            cell.value = helpers.modDateTime(datetime.datetime.now())
            cell.style.font.name = 'Arial'
            cell.style.font.size = 11
            cell.style.font.bold = True
            cell.style.alignment.wrap_text = False

            row_counter += 1
            row_counter += 1
            
            cols_count = len(kitchen_table['participants'])
            col_counter += 1
            self.setExcelHeaderCell(ws3, col_counter, row_counter, _(u"Total"), 4)
            col = get_column_letter(col_counter)

            cell = ws3.cell('%s%s'%(col, row_counter))
            cell.style.alignment.horizontal = cell.style.alignment.HORIZONTAL_CENTER
            cell.style.alignment.text_rotation = 90
            cell.style.alignment.wrap_text = True
            col_counter += 1
            for participant in kitchen_table['participants']:
                self.setExcelHeaderCell(ws3, col_counter, row_counter, kitchen_table['participants'][participant], 4)
                col = get_column_letter(col_counter)

                cell = ws3.cell('%s%s'%(col, row_counter))
                cell.style.alignment.horizontal = cell.style.alignment.HORIZONTAL_CENTER
                cell.style.alignment.text_rotation = 90
                cell.style.alignment.wrap_text = True
                col_counter += 1

            col_counter = 1
            row_counter += 1
            if len(kitchen_table['diets'])>0:
                self.setExcelHeaderCell(ws3, col_counter, row_counter, _('Diets'), 20)
                row_counter += 1
                col_counter = 1
                for diet in kitchen_table['diets']:
                    tmp_count = 0
                    self.setExcelCell(ws3, col_counter, row_counter, diet)
                    col_counter = 3
                    for participant in kitchen_table['participants']:
                        if participant in kitchen_table['diets'][diet]:
                            tmp_count += 1
                            self.setExcelCell(ws3, col_counter, row_counter, 'X')
                            self.setExcelCellFill(ws3, col_counter, row_counter, openpyxl_style.Color.GREEN)
                            col = get_column_letter(col_counter)

                            cell = ws3.cell('%s%s'%(col, row_counter))
                            cell.style.alignment.horizontal = cell.style.alignment.HORIZONTAL_CENTER
                            cell.style.font.bold = True

                        col_counter += 1
                    self.setExcelCell(ws3, 2, row_counter, tmp_count)
                    row_counter += 1
                    col_counter = 1
                row_counter += 1
                col_counter = 1
            
            if len(kitchen_table['food_allergies'])>0:
                self.setExcelHeaderCell(ws3, col_counter, row_counter, _('Food allergies'), 20)
                row_counter += 1
                col_counter = 1
                for food_allergy in kitchen_table['food_allergies']:
                    self.setExcelCell(ws3, col_counter, row_counter, food_allergy)
                    tmp_count = 0
                    col_counter = 3
                    for participant in kitchen_table['participants']:
                        if participant in kitchen_table['food_allergies'][food_allergy]:
                            tmp_count += 1
                            self.setExcelCell(ws3, col_counter, row_counter, 'X')
                            self.setExcelCellFill(ws3, col_counter, row_counter, openpyxl_style.Color.GREEN)
                            col = get_column_letter(col_counter)

                            cell = ws3.cell('%s%s'%(col, row_counter))
                            cell.style.alignment.horizontal = cell.style.alignment.HORIZONTAL_CENTER
                            cell.style.font.bold = True
                        col_counter += 1
                    self.setExcelCell(ws3, 2, row_counter, tmp_count)
                    row_counter += 1
                    col_counter = 1
                row_counter += 1
                col_counter = 1
        
        full_filename = base_path + dest_filename

        wb.save(filename = full_filename)
        self.request.response.content_type = "application/ms-excel"
        self.request.response.content_disposition = 'attachment; filename="leirirekkari_report_'+datetime.datetime.now().strftime("%Y%m%d_%H%M")+'.xls"'
        return FileResponse(full_filename, request=self.request)
    
    def getReportResults(self, request):
        
        filters = []

        filter_strings = {
        }


        to_view = {
            'basic_info': True,
            'booking_no': False,
            'polku_member_id': False,
            'status': True,
            'status_all': False,
            'age_group': True,
            'sex': False,
            'age': False,
            'birthdate': False,
            'club': True,
            'subunit': True,
            'village': True,
            'village_kitchen': False,
            'subcamp': True,
            'presence': True,
            'next_of_kin': False,
            'wishes': False,
            'address': False,
            'phone_and_email': False,
            'languages': False,
            'designation_title': False,
            'designation_all': False,
            'spiritual': False,
            'payments': False,
        }

        if security.has_permission('office_participant_view_medical', request.context, request):
            medical_to_view = {
                'medical_diets': False,
                'medical_diets_boolean': False,
                'medical_food_allergies': False,
                'medical_food_allergies_boolean': False,
                'medical_allergies': False,
                'medical_allergies_boolean': False,
                'medical_other': False,
                'medical_need_help_boolean': False,
                'kitchen_table': False,
            }

            to_view = dict(to_view.items() + medical_to_view.items())
        elif security.has_permission('kitchen_view_allergies', request.context, request):
            medical_to_view = {
                'medical_diets': False,
                'medical_diets_boolean': False,
                'medical_food_allergies': False,
                'medical_food_allergies_boolean': False,
                'medical_allergies': False,
                'medical_allergies_boolean': False,
                'kitchen_table': False,
            }
            to_view = dict(to_view.items() + medical_to_view.items())




        if len(request.POST.getall('to_view'))>0:
            to_view_items = request.POST.getall('to_view')
            for item in to_view:
                if item in to_view_items:
                    to_view[item] = True
                elif item in to_view:
                    to_view[item] = False

        show_canceled = request.POST.get('search_canceled')
        if show_canceled == 'Only':
            filters.append(Participant.active == False)
            filter_strings['show_canceled'] = show_canceled
        elif show_canceled == 'Both':
            filter_strings['show_canceled'] = show_canceled
        else:
            filters.append(Participant.active == True)
            filter_strings['show_canceled'] = 'No'

        show_open_payments = request.POST.get('search_open_payments')
        if show_open_payments == 'Only':
            not_paid = DBSession.query(ParticipantPayment).filter(ParticipantPayment.paid == False).all()
            not_paid_ids = []
            if len(not_paid) > 0:
                for tmp in not_paid:
                    not_paid_ids.append(tmp.participant_id)

            filters.append(Participant.id.in_(not_paid_ids))
            filter_strings['search_open_payments'] = show_open_payments

        if len(request.POST.getall('search_agegroup'))>0:
            search_agegroup = request.POST.getall('search_agegroup')
            filters.append(Participant.age_group.in_(search_agegroup))
            filter_strings['search_agegroup'] = search_agegroup
        if len(request.POST.getall('search_sex'))>0:
            search_sex = request.POST.getall('search_sex')
            filters.append(Participant.sex.in_(search_sex))
            filter_strings['search_sex'] = search_sex
        if len(request.POST.getall('search_subcamp'))>0:
            search_subcamp = request.POST.getall('search_subcamp')
            filters.append(Participant.subcamp_id.in_(search_subcamp))
            filter_strings['search_subcamp'] = search_subcamp
        if len(request.POST.getall('search_village'))>0:
            search_village = request.POST.getall('search_village')
            filters.append(Participant.village_id.in_(search_village))
            filter_strings['search_village'] = search_village
        if len(request.POST.getall('search_subunit'))>0:
            search_subunit = request.POST.getall('search_subunit')
            filters.append(Participant.subunit_id.in_(search_subunit))
            filter_strings['search_subunit'] = search_subunit
        if len(request.POST.getall('search_club'))>0:
            search_club = request.POST.getall('search_club')
            filters.append(Participant.club_id.in_(search_club))
            filter_strings['search_club'] = search_club
        if len(request.POST.getall('search_spiritual'))>0:
            search_spiritual = request.POST.getall('search_spiritual')
            filters.append(Participant.spiritual.in_(search_spiritual))
            filter_strings['search_spiritual'] = search_spiritual
        if len(request.POST.getall('search_village_kitchen'))>0:
            search_village_kitchen = request.POST.getall('search_village_kitchen')
            villages_ids = []
            for village_kitchen in search_village_kitchen:
                village_kitchen = organizationhelpers.getVillageKitchen(int(village_kitchen))
                if len(village_kitchen.villages) > 0:
                    for village in village_kitchen.villages:
                        villages_ids.append(village.id)
            filters.append(Participant.village_id.in_(villages_ids))
            filter_strings['search_village_kitchen'] = search_village_kitchen

        filter_strings['search_additional'] = []

        report = DBSession.query(Participant)

        if len(request.POST.getall('additional_search_row_field'))>0:
            
            fields = request.POST.getall('additional_search_row_field')
            types = request.POST.getall('additional_search_row_type')
            values = request.POST.getall('additional_search_row_value')
            method_map = {'==': '__eq__', '!=': '__ne__'}
            for key, value in enumerate(fields):
                if fields[key] != '':
                    if 'Participant.' in fields[key]:
                        if types[key] in ['==', '!=']:
                            comparison = getattr(eval(fields[key]), method_map[types[key]])
                            filters.append(comparison(values[key]))
                            filter_strings['search_additional'].append({'field':fields[key], 'type':types[key], 'value':values[key]})
                        elif types[key] == 'IN':
                            values = [x.strip() for x in values[key].split(',')]
                            filters.append(eval(fields[key]).in_(values))
                            filter_strings['search_additional'].append({'field':fields[key], 'type':types[key], 'value':values[key]})
                        elif types[key] == 'LIKE':
                            filters.append(eval(fields[key]).like('%'+values[key]+'%'))
                            filter_strings['search_additional'].append({'field':fields[key], 'type':types[key], 'value':values[key]})
                    else:
                        split_field = fields[key].split('.')
                        print split_field
                        report = report.join(eval(split_field[0]))
                        if split_field[0] == 'ParticipantNextOfKin':
                            field_key1 = split_field[0]+'.primary_'+split_field[1]
                            field_key2 = split_field[0]+'.secondary_'+split_field[1]
                            if types[key] in ['==', '!=']:
                                comparison1 = getattr(eval(field_key1), method_map[types[key]])
                                comparison2 = getattr(eval(field_key2), method_map[types[key]])
                                filters.append(or_(comparison1(values[key]), comparison2(values[key])))
                                filter_strings['search_additional'].append({'field':fields[key], 'type':types[key], 'value':values[key]})
                            elif types[key] == 'IN':
                                values = [x.strip() for x in values[key].split(',')]
                                filters.append(or_(eval(field_key1).in_(values), eval(field_key2).in_(values)))
                                filter_strings['search_additional'].append({'field':fields[key], 'type':types[key], 'value':values[key]})
                            elif types[key] == 'LIKE':
                                filters.append(or_(eval(field_key1).like('%'+values[key]+'%'), eval(field_key2).like('%'+values[key]+'%')))
                                filter_strings['search_additional'].append({'field':fields[key], 'type':types[key], 'value':values[key]})
                        
                        else:
                            if types[key] in ['==', '!=']:
                                comparison = getattr(eval(fields[key]), method_map[types[key]])
                                filters.append(comparison(values[key]))
                                filter_strings['search_additional'].append({'field':fields[key], 'type':types[key], 'value':values[key]})
                            elif types[key] == 'IN':
                                values = [x.strip() for x in values[key].split(',')]
                                filters.append(eval(fields[key]).in_(values))
                                filter_strings['search_additional'].append({'field':fields[key], 'type':types[key], 'value':values[key]})
                            elif types[key] == 'LIKE':
                                filters.append(eval(fields[key]).like('%'+values[key]+'%'))
                                filter_strings['search_additional'].append({'field':fields[key], 'type':types[key], 'value':values[key]})
                        


        if request.POST.get('search_birthdate') != None and request.POST.get('search_birthdate') != '':
            search_birthdate = request.POST.get('search_birthdate').strip()
            dt = helpers.parseFinnishDateFromString(search_birthdate)
            start_day = dt.timetuple().tm_mday
            start_month = dt.timetuple().tm_mon

            orm_start_day = sql.expression.extract('DAY', Participant.birthdate)
            orm_start_month = sql.expression.extract('MONTH', Participant.birthdate)

            filters.append(start_day == orm_start_day)
            filters.append(start_month == orm_start_month)
            filter_strings['search_birthdate'] = search_birthdate
            to_view['birthdate'] = True

        if len(request.POST.getall('search_status'))>0:
            search_status = request.POST.getall('search_status')
            filters.append(Participant.latest_status_key.in_(search_status))
            filter_strings['search_status'] = search_status


        if len(filters) > 0:
            report = report.filter(and_(*filters))
            

        if request.POST.get('search_datetime') != None and request.POST.get('search_datetime') != '':
            search_datetime = request.POST.get('search_datetime').strip()
            dt = helpers.parseFinnishDateFromString(search_datetime)
            report = report.join(ParticipantPresence).filter(ParticipantPresence.presence_starts <= dt, ParticipantPresence.presence_ends >= dt)
            filter_strings['search_datetime'] = search_datetime
            to_view['presence'] = True

        participants = report.all()
        
        return {'participants':participants, 'to_view':to_view, 'filter_strings':filter_strings}
    
    @view_config(route_name='office_presences_frontpage', renderer='office/presences/index.mak', permission='office_view')
    def office_presences_frontpage(self):
        if self.request.redirect_forbidden:
            return HTTPFound(location='/forbidden/')
        elif userhelpers.checkUserPasswordChangeNeed(self.request):
            return HTTPFound(location='/settings/me/edit/')
        _ = self.request.translate


        if self.request.method == 'POST' and self.request.POST.get('date') != None and self.request.POST.get('date').strip() != '':
            day = helpers.parseFinnishDateFromString(self.request.POST.get('date').strip())
        else:
            day = datetime.date.today()
        subcamps = DBSession.query(Subcamp).all()

        first_day = DBSession.query(Setting).with_entities(Setting.setting_key, Setting.setting_value).filter(Setting.setting_key == 'camp_first_day').first()
        last_day = DBSession.query(Setting).with_entities(Setting.setting_key, Setting.setting_value).filter(Setting.setting_key == 'camp_last_day').first()
        eating_times_breakfast = DBSession.query(Setting).with_entities(Setting.setting_key, Setting.setting_value).filter(Setting.setting_key == 'eating_times_breakfast').first()
        eating_times_lunch = DBSession.query(Setting).with_entities(Setting.setting_key, Setting.setting_value).filter(Setting.setting_key == 'eating_times_lunch').first()
        eating_times_dinner = DBSession.query(Setting).with_entities(Setting.setting_key, Setting.setting_value).filter(Setting.setting_key == 'eating_times_dinner').first()
        eating_times_supper = DBSession.query(Setting).with_entities(Setting.setting_key, Setting.setting_value).filter(Setting.setting_key == 'eating_times_supper').first()
        start_dt = helpers.parseFinnishDateFromString(first_day[1])
        end_dt = helpers.parseFinnishDateFromString(last_day[1])

        day_count = (end_dt - start_dt).days + 1

        dates = []
        dates.append(day)

        days = {
            'start_dt':start_dt,
            'end_dt':end_dt,
            'dates':dates
        }



        eating_times = {
            'eating_times_breakfast':eating_times_breakfast[1],
            'eating_times_lunch':eating_times_lunch[1],
            'eating_times_dinner':eating_times_dinner[1],
            'eating_times_supper':eating_times_supper[1],
        }

        self.request.bread.append({'url':'/office/', 'text':_('Office')})
        self.request.bread.append({'url':'/office/presences/', 'text':_('Presences')})
        return {'day':day,'subcamps':subcamps, 'days':days, 'eating_times':eating_times}

    @view_config(route_name='office_presences_clubs', renderer='office/presences/clubs.mak', permission='office_view')
    def office_presences_clubs(self):
        if self.request.redirect_forbidden:
            return HTTPFound(location='/forbidden/')
        elif userhelpers.checkUserPasswordChangeNeed(self.request):
            return HTTPFound(location='/settings/me/edit/')
        _ = self.request.translate


        if self.request.method == 'POST' and self.request.POST.get('date') != None and self.request.POST.get('date').strip() != '':
            day = helpers.parseFinnishDateFromString(self.request.POST.get('date').strip())
        else:
            day = datetime.date.today()
        subcamps = DBSession.query(Subcamp).all()

        first_day = DBSession.query(Setting).with_entities(Setting.setting_key, Setting.setting_value).filter(Setting.setting_key == 'camp_first_day').first()
        last_day = DBSession.query(Setting).with_entities(Setting.setting_key, Setting.setting_value).filter(Setting.setting_key == 'camp_last_day').first()
        eating_times_breakfast = DBSession.query(Setting).with_entities(Setting.setting_key, Setting.setting_value).filter(Setting.setting_key == 'eating_times_breakfast').first()
        eating_times_lunch = DBSession.query(Setting).with_entities(Setting.setting_key, Setting.setting_value).filter(Setting.setting_key == 'eating_times_lunch').first()
        eating_times_dinner = DBSession.query(Setting).with_entities(Setting.setting_key, Setting.setting_value).filter(Setting.setting_key == 'eating_times_dinner').first()
        eating_times_supper = DBSession.query(Setting).with_entities(Setting.setting_key, Setting.setting_value).filter(Setting.setting_key == 'eating_times_supper').first()
        start_dt = helpers.parseFinnishDateFromString(first_day[1])
        end_dt = helpers.parseFinnishDateFromString(last_day[1])

        day_count = (end_dt - start_dt).days + 1

        dates = []
        dates.append(day)

        days = {
            'start_dt':start_dt,
            'end_dt':end_dt,
            'dates':dates
        }



        eating_times = {
            'eating_times_breakfast':eating_times_breakfast[1],
            'eating_times_lunch':eating_times_lunch[1],
            'eating_times_dinner':eating_times_dinner[1],
            'eating_times_supper':eating_times_supper[1],
        }

        self.request.bread.append({'url':'/office/', 'text':_('Office')})
        self.request.bread.append({'url':'/office/presences/', 'text':_('Presences')})
        return {'day':day,'subcamps':subcamps, 'days':days, 'eating_times':eating_times}

        
def includeme(config):
    config.add_route('office_frontpage', '/office/')
    config.add_route('office_participant_new', '/office/participant/new/')
    config.add_route('office_participant_view', '/office/participant/view/{participant_id}/')
    config.add_route('office_participant_view_payments', '/office/participant/view/{participant_id}/payments/')
    config.add_route('office_participant_cancel', '/office/participant/view/{participant_id}/cancel/')
    config.add_route('office_participant_uncancel', '/office/participant/view/{participant_id}/uncancel/')
    config.add_route('office_participant_edit', '/office/participant/edit/{participant_id}/')
    config.add_route('office_participant_search', '/office/participant/search/')
    config.add_route('office_report', '/office/report/')
    config.add_route('office_report_excel', '/office/report_excel/')
    config.add_route('office_presences_frontpage', '/office/presences/')
    config.add_route('office_presences_clubs', '/office/presences/clubs/')